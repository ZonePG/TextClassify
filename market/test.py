import torch
import torch.nn as nn
import numpy as np
import torch.nn.functional as F
from utils import build_iterator, build_vocab


class Config(object):

    """配置参数"""

    def __init__(self, dataset, embedding):
        self.model_name = "TextCNN"
        self.train_path = dataset + "/data/train.txt"  # 训练集
        self.dev_path = dataset + "/data/dev.txt"  # 验证集
        self.test_path = dataset + "/data/test.txt"  # 测试集
        self.class_list = [
            x.strip()
            for x in open(dataset + "/data/class.txt", encoding="utf-8").readlines()
        ]  # 类别名单
        self.vocab_path = dataset + "/data/vocab.pkl"  # 词表
        self.save_path = dataset + "/saved_dict/" + self.model_name + ".ckpt"  # 模型训练结果
        self.log_path = dataset + "/log/" + self.model_name
        self.embedding_pretrained = (
            torch.tensor(
                np.load(dataset + "/data/" + embedding)["embeddings"].astype("float32")
            )
            if embedding != "random"
            else None
        )  # 预训练词向量
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")  # 设备

        self.dropout = 0.5  # 随机失活
        self.require_improvement = 1000  # 若超过1000batch效果还没提升，则提前结束训练
        self.num_classes = len(self.class_list)  # 类别数
        self.n_vocab = 0  # 词表大小，在运行时赋值
        self.num_epochs = 20  # epoch数
        self.batch_size = 128  # mini-batch大小
        self.pad_size = 32  # 每句话处理成的长度(短填长切)
        self.learning_rate = 1e-3  # 学习率
        self.embed = (
            self.embedding_pretrained.size(1)
            if self.embedding_pretrained is not None
            else 300
        )  # 字向量维度
        self.filter_sizes = (2, 3, 4)  # 卷积核尺寸
        self.num_filters = 256  # 卷积核数量(channels数)


class Model(nn.Module):
    def __init__(self, config):
        super(Model, self).__init__()
        if config.embedding_pretrained is not None:
            self.embedding = nn.Embedding.from_pretrained(
                config.embedding_pretrained, freeze=False
            )
        else:
            self.embedding = nn.Embedding(
                config.n_vocab, config.embed, padding_idx=config.n_vocab - 1
            )
        self.convs = nn.ModuleList(
            [
                nn.Conv2d(1, config.num_filters, (k, config.embed))
                for k in config.filter_sizes
            ]
        )
        self.dropout = nn.Dropout(config.dropout)
        self.fc = nn.Linear(
            config.num_filters * len(config.filter_sizes), config.num_classes
        )

    def conv_and_pool(self, x, conv):
        x = F.relu(conv(x)).squeeze(3)
        x = F.max_pool1d(x, x.size(2)).squeeze(2)
        return x

    def forward(self, x):
        out = self.embedding(x[0])
        out = out.unsqueeze(1)
        out = torch.cat([self.conv_and_pool(out, conv) for conv in self.convs], 1)
        out = self.dropout(out)
        out = self.fc(out)
        return out


UNK, PAD = "<UNK>", "<PAD>"  # 未知字，padding符号


def convert(content):
    content = content.replace("\n", "")
    content = content.replace("\u3000", "")
    content = content.replace(" ", "")
    content = content.replace("\xa0", "")
    content = content.replace("\t", "")

    str2list = list(content)
    if len(str2list) <= 256:
        return content
    else:
        list2str = "".join(content[:256])
        return list2str


def load_dataset(data, config):
    pad_size = config.pad_size
    contents = []
    tokenizer = lambda x: [y for y in x]  # char-level
    import pickle as pkl
    import os

    MAX_VOCAB_SIZE = 10000  # 词表长度限制

    if os.path.exists(config.vocab_path):
        vocab = pkl.load(open(config.vocab_path, "rb"))
    else:
        vocab = build_vocab(
            config.train_path, tokenizer=tokenizer, max_size=MAX_VOCAB_SIZE, min_freq=1
        )
        pkl.dump(vocab, open(config.vocab_path, "wb"))
    for line in data:
        lin = convert(line)
        if not line:
            continue
        words_line = []
        token = tokenizer(lin)
        seq_len = len(token)
        if pad_size:
            if len(token) < pad_size:
                token.extend([PAD] * (pad_size - len(token)))
            else:
                token = token[:pad_size]
                seq_len = pad_size
        for word in token:
            words_line.append(vocab.get(word, vocab.get(UNK)))
        contents.append((words_line, int(0), seq_len))
    return contents


def match_label(pred, config):
    label_list = config.class_list
    return label_list[pred]


def final_predict(config, model, data_iter):
    map_location = lambda storage, loc: storage
    model.load_state_dict(torch.load(config.save_path, map_location=map_location))
    model.eval()
    predict_all = np.array([])
    with torch.no_grad():
        for texts, _ in data_iter:
            outputs = model(texts)
            pred = torch.max(outputs.data, 1)[1].cpu().numpy()
            pred_label = [match_label(i, config) for i in pred]
            predict_all = np.append(predict_all, pred_label)

    return predict_all


def predict(text):
    dataset = "datasets"  # 数据集
    # 搜狗新闻:embedding_SougouNews.npz, 腾讯:embedding_Tencent.npz, 随机初始化:random
    embedding = "embedding_SougouNews.npz"
    config = Config(dataset, embedding)
    model = Model(config).to(config.device)
    test_data = load_dataset(text, config)
    test_iter = build_iterator(test_data, config)
    result = final_predict(config, model, test_iter)
    for i, j in enumerate(result):
        print("text:{}".format(text[i]))
        print("label:{}".format(j))
        print()


def classify(text):
    text = [text]
    dataset = "market/datasets"  # 数据集
    # 搜狗新闻:embedding_SougouNews.npz, 腾讯:embedding_Tencent.npz, 随机初始化:random
    embedding = "embedding_SougouNews.npz"
    config = Config(dataset, embedding)
    model = Model(config).to(config.device)
    test_data = load_dataset(text, config)
    test_iter = build_iterator(test_data, config)
    result = final_predict(config, model, test_iter)
    # print('text:{}'.format(text[i]))
    # print('label:{}'.format(j))
    return result[0]


def predict_test(text):
    res = []

    dataset = "datasets"  # 数据集
    # 搜狗新闻:embedding_SougouNews.npz, 腾讯:embedding_Tencent.npz, 随机初始化:random
    embedding = "embedding_SougouNews.npz"
    config = Config(dataset, embedding)
    model = Model(config).to(config.device)
    test_data = load_dataset(text, config)
    test_iter = build_iterator(test_data, config)
    result = final_predict(config, model, test_iter)
    for i, j in enumerate(result):
        # print ("text:{}".format(text[i]))
        # print ("label:{}".format(j))
        # print ()
        res.append(j)

    return res


if __name__ == "__main__":
    # test = ['国考28日网上查报名序号查询后务必牢记报名参加2011年国家公务员的考生，如果您已通过资格审查，那么请于10月28日8：00后，登录考录专题网站查询自己的“关键数字”——报名序号。'
    #     '国家公务员局等部门提醒：报名序号是报考人员报名确认和下载打印准考证等事项的重要依据和关键字，请务必牢记。此外，由于年龄在35周岁以上、40周岁以下的应届毕业硕士研究生和'
    #     '博士研究生(非在职)，不通过网络进行报名，所以，这类人报名须直接与要报考的招录机关联系，通过电话传真或发送电子邮件等方式报名。',
    #     '高品质低价格东芝L315双核本3999元作者：徐彬【北京行情】2月20日东芝SatelliteL300(参数图片文章评论)采用14.1英寸WXGA宽屏幕设计，配备了IntelPentiumDual-CoreT2390'
    #     '双核处理器(1.86GHz主频/1MB二级缓存/533MHz前端总线)、IntelGM965芯片组、1GBDDR2内存、120GB硬盘、DVD刻录光驱和IntelGMAX3100集成显卡。目前，它的经销商报价为3999元。',
    #     '国安少帅曾两度出山救危局他已托起京师一代才俊新浪体育讯随着联赛中的连续不胜，卫冕冠军北京国安的队员心里到了崩溃的边缘，俱乐部董事会连夜开会做出了更换主教练洪元硕的决定。'
    #     '而接替洪元硕的，正是上赛季在李章洙下课风波中同样下课的国安俱乐部副总魏克兴。生于1963年的魏克兴球员时代并没有特别辉煌的履历，但也绝对称得上特别：15岁在北京青年队获青年'
    #     '联赛最佳射手，22岁进入国家队，著名的5-19一战中，他是国家队的替补队员。',
    #     '汤盈盈撞人心情未平复眼泛泪光拒谈悔意(附图)新浪娱乐讯汤盈盈日前醉驾撞车伤人被捕，原本要彩排《欢乐满东华2008》的她因而缺席，直至昨日(12月2日)，盈盈继续要与王君馨、马'
    #     '赛、胡定欣等彩排，大批记者在电视城守候，她足足迟了约1小时才到场。全身黑衣打扮的盈盈，神情落寞、木无表情，回答记者问题时更眼泛泪光。盈盈因为迟到，向记者说声“不好意思”后'
    #     '便急步入场，其助手坦言盈盈没什么可以讲。后来在《欢乐满东华2008》监制何小慧陪同下，盈盈接受简短访问，她小声地说：“多谢大家关心，交给警方处理了，不方便讲，',
    #     '甲醇期货今日挂牌上市继上半年焦炭、铅期货上市后，酝酿已久的甲醇期货将在今日正式挂牌交易。基准价均为3050元／吨继上半年焦炭、铅期货上市后，酝酿已久的甲醇期货将在今日正式'
    #     '挂牌交易。郑州商品交易所（郑商所）昨日公布首批甲醇期货8合约的上市挂牌基准价，均为3050元／吨。据此推算，买卖一手甲醇合约至少需要12200元。业内人士认为，作为国际市场上的'
    #     '首个甲醇期货品种，其今日挂牌后可能会因炒新资金追捧而出现冲高走势，脉冲式行情过后可能有所回落，不过，投资者在上市初期应关注期现价差异常带来的无风险套利交易机会。',
    #     '原标题：力破“四唯”，建立“多维”科研评价体系“论文、职称、学历、奖项”是科技人才评价体系中的重要指标。很长一段时间，以此四项为导向的科技人才评价体系推动了中国科学技术的发展。但是，唯此四项的机械作法，却逐渐发展演变为科研工作的羁绊和科研人员的包袱，饱受诟病。近年来，我国持续深化科技体制改革，完善科研评价机制，力破“四唯”。',
    #     '天津建立培育机制助力留学人员企业创新发展',
    #     '人民时评：以职业教育赋能脱贫攻坚',
    #     '《金牌喜剧班》神秘开场学员直面导师灵魂拷问 “班级大考”谁是人气No.1',
    #     '题电影系列项目包括一部献礼院线电影《我和我的奥运》和多部奥运城市影片《我的奥运，我的城》。启动仪式现场。活扬光大，进而对中国人的生活、工作、城市新时代精神面貌进行全景式呈现，传递奥运精神的同时，传播新时代中国声音。']
    # predict(test)

    xlsx = []
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter

    # READ
    filename = input("Please input xlsx file name(example: test_data.xlsx): ")
    # 打开一个workbook
    print("loading data....")
    wb = load_workbook(filename)

    # 获取所有表格(worksheet)的名字
    sheets = wb.sheetnames
    # 第一个表格的名称
    sheet_first = sheets[0]
    # 获取特定的worksheet
    ws = wb[sheet_first]

    # 获取表格所有行和列，两者都是可迭代的
    rows = ws.rows
    columns = ws.columns

    # 迭代所有的行
    for row in rows:
        line = [col.value for col in row]
        xlsx.append(line)

    # # 通过坐标读取值
    # print ws.cell("A1").value  # A表示列,1表示行
    # print ws.cell(row=1, column=1).value
    # print(xlsx[1])
    # print(len(xlsx))

    text = [line[2] + '\n' + line[3] for line in xlsx]
    text = text[1:]
    # print(text[0])

    print("text classifing....")
    label = predict_test(text)
    # print(label[0])

    # Write
    # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
    wb = Workbook()

    # 获取当前活跃的worksheet,默认就是第一个worksheet
    ws = wb.active

    print("result saving...")
    # 设置单元格的值，A1等于6(测试可知openpyxl的行和列编号从1开始计算)，B1等于7
    ws.cell(row=1, column=1).value = "编号"
    ws.cell(row=1, column=2).value = "ChannelName"
    ws.cell(row=1, column=3).value = "title"
    ws.cell(row=1, column=4).value = "content"

    for row in range(1, len(xlsx)):
        ws.cell(row=row+1, column=1).value = xlsx[row][0]
        ws.cell(row=row+1, column=2).value = label[row-1]
        ws.cell(row=row+1, column=3).value = xlsx[row][2]
        ws.cell(row=row+1, column=4).value = xlsx[row][3]

    # 保存
    wb.save(filename="output.xlsx")
    print("finish! result saved in output.xlsx!!!")
